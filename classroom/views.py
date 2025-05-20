from django.shortcuts import render, get_object_or_404, redirect
from rest_framework.viewsets import ModelViewSet
from .serializers import TeacherAttendanceSerializer, ClassroomModeSerializer
import random
from subject.models import Subject, Schedule
from course.models import SubjectEnrollment, Semester
from accounts.models import CustomUser
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import action
from .models import *
from rest_framework.response import Response
from django.utils.timezone import localtime, now
from django.db.models import Sum, F, ExpressionWrapper, DurationField
from django.db.models.functions import TruncDate
from rest_framework import status
from datetime import datetime, timedelta
from calendars.models import Holiday
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from django.core.files.base import ContentFile
import base64
import json
from django.contrib.auth.decorators import login_required
import calendar
from django.db.models import OuterRef, Subquery, Value, CharField
from django.db.models.functions import Coalesce
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import io
# Create your views here.

def format_total_time(duration):
    if duration is None:
        return "N/A"
    total_seconds = duration.total_seconds()
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    return f"{hours}h {minutes}m"

class TeacherAttendanceViewSet(ModelViewSet):
    serializer_class = TeacherAttendanceSerializer

    def get_queryset(self):
        return self.serializer_class.Meta.model.objects.all()
    
    def list(self, request, *args, **kwargs):
        view_type = request.query_params.get('view_type', 'daily')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        today = datetime.now().date()
        if view_type == 'daily':
            if not start_date:
                start_date = today
            if not end_date:
                end_date = start_date
        elif view_type == 'weekly':
            if not start_date:
                start_date = today - timedelta(days=today.weekday())
            if not end_date:
                end_date = start_date + timedelta(days=6)
        elif view_type == 'monthly':
            if not start_date:
                start_date = today.replace(day=1)
            if not end_date:
                end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)

        # Generate the full date range
        date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

        # Fetch holidays in the date range
        holidays = Holiday.objects.filter(date__gte=start_date, date__lte=end_date)
        holiday_map = {holiday.date: (holiday.title, holiday.holiday_type) for holiday in holidays}

        # Fetch attendance data
        attendance_data = Teacher_Attendance.objects.filter(
            time_started__date__gte=start_date,
            time_started__date__lte=end_date,
        ).annotate(
            day=TruncDate('time_started'),
            duration=ExpressionWrapper(F('time_ended') - F('time_started'), output_field=DurationField()),
        )

        attendance_map = {
            (att.teacher_id, att.subject_id, att.day): att for att in attendance_data
        }

        # Fetch schedules
        schedules = Schedule.objects.prefetch_related('subject__assign_teacher', 'subject__substitute_teacher')

        grouped_data = {}
        for schedule in schedules:
            subject = schedule.subject
            teacher = subject.active_teacher
            if not teacher:
                continue

            teacher_name = f"{teacher.first_name} {teacher.last_name}"
            subject_name = subject.subject_name

            if teacher_name not in grouped_data:
                grouped_data[teacher_name] = {}

            if subject_name not in grouped_data[teacher_name]:
                grouped_data[teacher_name][subject_name] = []


            schedule_dates = [
                current_date for current_date in date_range
                if current_date.strftime('%a') in schedule.days_of_week
            ]

            for current_date in date_range:
                is_holiday = current_date in holiday_map
                attendance = attendance_map.get((teacher.id, subject.id, current_date))
                today = datetime.now().date()

                attendance_id = None 
                screenshots = []
                if attendance:
                    screenshots = Screenshot.objects.filter(teacher_attendance=attendance).values_list('image', flat=True)
                    attendance_id = attendance.id

                if is_holiday:
                    holiday_title, holiday_type = holiday_map[current_date]
                    if schedule.schedule_type in ['Regular', 'Build in']:
                        # Calculate total time for the scheduled period
                        schedule_start = datetime.combine(current_date, schedule.schedule_start_time)
                        schedule_end = datetime.combine(current_date, schedule.schedule_end_time)
                        total_time = format_total_time(schedule_end - schedule_start)
                    else:
                        total_time = "N/A"

                    grouped_data[teacher_name][subject_name].append({
                        "attendance_id": attendance_id,
                        "date": current_date,
                        "schedule": f"{schedule.schedule_start_time} - {schedule.schedule_end_time}",
                        "schedule_type": schedule.schedule_type,
                        "status": f"Holiday ({holiday_type}) - {holiday_title}",
                        "time_started": None,
                        "time_ended": None,
                        "total_time": total_time,
                        "screenshots": list(screenshots),
                    })
                elif current_date in schedule_dates:
                    if attendance:
                        grouped_data[teacher_name][subject_name].append({
                            "attendance_id": attendance_id,
                            "date": current_date,
                            "schedule": f"{schedule.schedule_start_time} - {schedule.schedule_end_time}",
                            "schedule_type": schedule.schedule_type,
                            "status": "Present",
                            "time_started": attendance.time_started,
                            "time_ended": attendance.time_ended,
                            "total_time": format_total_time(attendance.duration),
                            "screenshots": list(screenshots),
                        })
                    else:
                        if current_date < today:
                            # Past scheduled day with no attendance recorded
                            grouped_data[teacher_name][subject_name].append({
                                "attendance_id": attendance_id,
                                "date": current_date,
                                "schedule": f"{schedule.schedule_start_time} - {schedule.schedule_end_time}",
                                "schedule_type": schedule.schedule_type,
                                "status": "Absent",
                                "time_started": None,
                                "time_ended": None,
                                "total_time": "N/A",
                                "screenshots": [],
                            })
                        else:
                            # Future scheduled day with no attendance
                            grouped_data[teacher_name][subject_name].append({
                                "attendance_id": attendance_id,
                                "date": current_date,
                                "schedule": f"{schedule.schedule_start_time} - {schedule.schedule_end_time}",
                                "schedule_type": schedule.schedule_type,
                                "status": "No Class",
                                "time_started": None,
                                "time_ended": None,
                                "total_time": "N/A",
                                "screenshots": [],
                            })
                else:
                    # No schedule for this date
                    grouped_data[teacher_name][subject_name].append({
                        "attendance_id": attendance_id,
                        "date": current_date,
                        "schedule": "No Schedule",
                        "schedule_type": schedule.schedule_type,
                        "status": "No Schedule",
                        "time_started": None,
                        "time_ended": None,
                        "total_time": "N/A",
                        "screenshots": [],
                    })

        return Response(grouped_data)
    
    def export_to_excel(self, request, *args, **kwargs):
        current_date = datetime.now().date()

        # Determine the current semester
        current_semester = Semester.objects.filter(start_date__lte=current_date, end_date__gte=current_date).first()
        current_semester_name = current_semester.semester_name if current_semester else "N/A"

        view_type = request.query_params.get('view_type', 'daily')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Call the existing list method to get the grouped data
        response = self.list(request, *args, **kwargs)
        grouped_data = response.data

        # Create an Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Teacher Attendance"

        # Write headers
        headers = [
            "Teacher Name", "Subject Name", "Subject Code", "Current Semester",
            "Date", "Schedule", "Schedule Type", "Status", "Time Started", "Time Ended", "Total Time"
        ]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Populate data
        row_num = 2
        for teacher, subjects in grouped_data.items():
            for subject, attendance_list in subjects.items():
                
                subject_obj = Subject.objects.filter(subject_name=subject).first()
                subject_code = subject_obj.subject_code if subject_obj else "N/A"

                for attendance in attendance_list:
                    # Ensure datetime fields are naive
                    date = attendance["date"]
                    time_started = attendance["time_started"]
                    time_ended = attendance["time_ended"]

                    # Convert datetime values to naive (remove timezone)
                    if isinstance(date, datetime):
                        date = date.replace(tzinfo=None).strftime('%Y-%m-%d')
                    if isinstance(time_started, datetime):
                        time_started = time_started.replace(tzinfo=None).strftime('%H:%M:%S')
                    if isinstance(time_ended, datetime):
                        time_ended = time_ended.replace(tzinfo=None).strftime('%H:%M:%S')

                    sheet.cell(row=row_num, column=1, value=teacher)
                    sheet.cell(row=row_num, column=2, value=subject)
                    sheet.cell(row=row_num, column=3, value=subject_code)
                    sheet.cell(row=row_num, column=4, value=current_semester_name)
                    sheet.cell(row=row_num, column=5, value=date)
                    sheet.cell(row=row_num, column=6, value=attendance["schedule"])
                    sheet.cell(row=row_num, column=7, value=attendance["schedule_type"])
                    sheet.cell(row=row_num, column=8, value=attendance["status"])
                    sheet.cell(row=row_num, column=9, value=time_started)
                    sheet.cell(row=row_num, column=10, value=time_ended)
                    sheet.cell(row=row_num, column=11, value=attendance["total_time"])
                    row_num += 1

        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = 20

        # Create HTTP response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename=Teacher_Attendance_{view_type}.xlsx"
        workbook.save(response)
        return response

    @action(detail=True, methods=['post'], url_path='start-class')
    def start_class(self, request, pk=None):
        subject_id = pk
        user = request.user

        # Ensure the subject exists (add your Subject model import)
        try:
            subject = Subject.objects.get(pk=subject_id)
        except Subject.DoesNotExist:
            return Response({'error': 'Subject not found'}, status=404)

        # Check if the user is the assigned teacher or substitute teacher
        if subject.assign_teacher != user and not (subject.substitute_teacher == user and subject.allow_substitute_teacher):
            return Response({'error': 'You are not assigned to this subject.'}, status=403)

        # Validate if the current time is within the schedule
        current_time = localtime(now())
        current_day = current_time.strftime('%a')
        

        schedules = subject.schedules.filter(days_of_week__icontains=current_day)
        if not schedules.exists():
            return Response({'error': f'No schedules found for today ({current_day}).'}, status=400)
        
        is_within_schedule = any(schedule.schedule_start_time <= current_time.time() <= schedule.schedule_end_time for schedule in schedules)

        if not is_within_schedule:
            return Response({'error': f'You can only start the class during your scheduled time.'}, status=400)


        # Create the teacher attendance record
        teacher_attendance = Teacher_Attendance.objects.create(
            subject=subject,
            teacher=user,
            time_started=current_time,
            is_active=True
        )

        serializer = self.get_serializer(teacher_attendance)
        return Response({
            'message': 'Class started successfully!',
            'attendance_id': teacher_attendance.id,
            'data': serializer.data
        }, status=201)
    

    @action(detail=True, methods=['post'], url_path='end-class')
    def end_class(self, request, pk=None):
        try:
            teacher_attendance = Teacher_Attendance.objects.get(subject__id=pk, teacher=request.user, is_active=True)
            teacher_attendance.time_ended = now()
            teacher_attendance.is_active = False
            teacher_attendance.save()
            return Response({'message': 'Classroom mode ended successfully.'}, status=200)
        except Teacher_Attendance.DoesNotExist:
            return Response({'error': 'Active classroom mode not found.'}, status=404)
    
        

    @action(detail=True, methods=['get'], url_path='current-state')
    def current_state(self, request, pk=None):
        try:
            teacher_attendance = Teacher_Attendance.objects.filter(
                subject__id=pk, teacher=request.user, is_active=True
            ).order_by('-time_started').first()
            if teacher_attendance:
                return Response({'is_active': True}, status=200)
            return Response({'is_active': False}, status=200)
        except Exception as e:
            return Response({'error': str(e)}, status=500)
        

    @action(detail=True, methods=['get'], url_path='get-end-time')
    def get_end_time(self, request, pk=None):
        """Fetches the scheduled end time for a class based on subject ID."""
        try:
            subject = Subject.objects.get(pk=pk)  # Fetch by subject ID
            current_day = localtime(now()).strftime('%a')

            # Get the latest end time for today
            schedules = subject.schedules.filter(days_of_week__icontains=current_day)
            if schedules.exists():
                latest_end_time = max(schedule.schedule_end_time for schedule in schedules)
                return Response({"end_time": latest_end_time.strftime('%H:%M:%S')}, status=200)

            return Response({"error": "No schedule found for today."}, status=404)

        except Subject.DoesNotExist:
            return Response({"error": "Subject not found."}, status=404)


class ClassroomModeViewSet(ModelViewSet):
    serializer_class = ClassroomModeSerializer

    def get_queryset(self):
        return self.serializer_class.Meta.model.objects.all()
    
    @action(detail=False, methods=['post'], url_path='toggle-mode')
    def toggle_classroom_mode(self, request):
        classroom_mode_instance, created = Classroom_mode.objects.get_or_create(id=1)  
        classroom_mode_instance.is_classroom_mode = not classroom_mode_instance.is_classroom_mode
        classroom_mode_instance.save()

        return Response(
            {"is_classroom_mode": classroom_mode_instance.is_classroom_mode},
            status=status.HTTP_200_OK
        )

@login_required
def enter_classroom_mode_view(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    classroom_mode_instance, _ = Classroom_mode.objects.get_or_create(subject=subject)
    classroom_mode_instance.is_classroom_mode = True
    classroom_mode_instance.save()
    return redirect('classroom_mode', pk=subject.id)

@login_required
def exit_classroom_mode_view(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    try:
        classroom_mode_instance = Classroom_mode.objects.get(subject=subject)
        classroom_mode_instance.is_classroom_mode = False
        classroom_mode_instance.save()
    except Classroom_mode.DoesNotExist:
        pass  # No need to do anything if the mode was never started
    return redirect('subjectDetail', pk=subject.id)


def lucky_draw(request, subject_id):
    """
    View for conducting a lucky draw for students enrolled in a specific subject
    within the current semester. Allows manual inclusion of a student and
    excluding previously winning students.
    """
    # Get the current semester based on today's date
    current_date = timezone.localtime(timezone.now()).date()
    current_semester = Semester.objects.filter(start_date__lte=current_date, end_date__gte=current_date).first()

    if not current_semester:
        return JsonResponse({
            'error': 'No active semester is found.',
            'students': [],
            'winner': None,
        }, status=404)

    # Get the subject
    subject = get_object_or_404(Subject, id=subject_id)

    # Get all students enrolled in the subject for the current semester
    students = list(SubjectEnrollment.objects.filter(
        subject=subject,
        semester=current_semester,
        status='enrolled'
    ).values('student__id', 'student__first_name', 'student__last_name'))

    # Optionally exclude previously winning students
    exclude_winners = request.GET.get('exclude_winners', 'false').lower() == 'true'
    if exclude_winners:
        # Fetch previously winning students (this can be fetched from a model or session)
        # For simplicity, assume `winning_students` is stored in the session
        winning_students = request.session.get('winning_students', [])
        students = [student for student in students if str(student['student__id']) not in winning_students]

    # Allow manual inclusion of a student
    manual_student_id = request.GET.get('manual_student_id', None)
    if manual_student_id:
        student = CustomUser.objects.filter(id=manual_student_id).values(
            'id', 'first_name', 'last_name'
        ).first()
        if student and student not in students:
            students.append({
                'student__id': student['id'],
                'student__first_name': student['first_name'],
                'student__last_name': student['last_name']
            })

    # Randomly select a winner
    winner = random.choice(students) if students else None

    # Save the winner to session if one exists
    if winner:
        winning_students = request.session.get('winning_students', [])
        winning_students.append(str(winner['student__id']))
        request.session['winning_students'] = winning_students

    return JsonResponse({
        'subject': subject.subject_name,
        'students': students,
        'winner': winner,
    })


def lucky_draw_page(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    return render(request, 'classroom/lucky_draw.html', {'subject': subject})

def classroom_dashboard(request):
    return render(request, 'classroom/Classroom_dashboard.html')

@csrf_exempt
def reset_lucky_draw(request, subject_id):
    """
    View to reset the lucky draw by clearing the list of previously winning students.
    """
    subject = get_object_or_404(Subject, id=subject_id)

    # Clear the session data for winning students
    if 'winning_students' in request.session:
        del request.session['winning_students']

    return JsonResponse({
        'message': f"The lucky draw for {subject.subject_name} has been reset.",
        'status': 'success'
    })

@login_required
def teacher_attendance(request):
    subject = Subject.objects.all()
    return render(request, 'teacher_attendance/teacher_attendance_list.html',{'subject': subject})


@login_required
def teacher_attendance_details(request, id):
    attendance_details = Teacher_Attendance.objects.filter(subject=id)
    
    subject_id = attendance_details.first().subject.id

    # Calculate total time for each attendance record
    for attendance in attendance_details:
        if attendance.time_ended:
            total_seconds = (attendance.time_ended - attendance.time_started).total_seconds()
            hours = int(total_seconds // 3600)
            minutes = int((total_seconds % 3600) // 60)
            attendance.total_time = f"{hours}h {minutes}m"
        else:
            attendance.total_time = "N/A"

    return render(request, 'teacher_attendance/teacher_attendance_details.html', {
        'attendance_details': attendance_details,
        "subject_id": subject_id,
    })


@login_required
def teacher_attendance_details_per_day(request, id):
    date_selected = request.GET.get("date")  # Get date from the URL

    # Validate date format
    try:
        selected_date = datetime.strptime(date_selected, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return HttpResponse("Invalid date format", status=400)

    # Filter attendance records based on subject and date
    attendance_details = Teacher_Attendance.objects.filter(
        subject_id=id,
        time_started__date=selected_date
    )

    # Calculate total time for each attendance record
    for attendance in attendance_details:
        if attendance.time_ended:
            total_seconds = (attendance.time_ended - attendance.time_started).total_seconds()
            hours, remainder = divmod(total_seconds, 3600)
            minutes, _ = divmod(remainder, 60)
            attendance.total_time = f"{hours}h {minutes}m"
        else:
            attendance.total_time = "N/A"

    return render(request, 'teacher_attendance/teacher_attendance_details.html', {
        'attendance_details': attendance_details,
        'selected_date': selected_date
    })



@login_required
def view_screenshots(request, id):
    attendance = get_object_or_404(Teacher_Attendance, id=id)
    screenshots = Screenshot.objects.filter(teacher_attendance=attendance)
    subject_id = attendance.subject.id if attendance.subject else None

    return render(request, "teacher_attendance/view_all_screenshot.html", {
        "attendance": attendance,
        "screenshots": screenshots,
        "subject_id": subject_id,
    })


@login_required
def view_screenshots_per_date(request, subject_id, selected_date):
    """
    View screenshots only for the given subject and date.
    """
    try:
        selected_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
        print(f"✅ Parsed selected_date: {selected_date}")  # ✅ Print parsed date

    except ValueError:
        print("❌ Invalid date format received!")  # ✅ Error Debugging
        return HttpResponse("Invalid date format", status=400)

    # ✅ Filter attendance records for the selected subject and date
    attendance_records = Teacher_Attendance.objects.filter(
        subject_id=subject_id,
        time_started__date=selected_date
    )

    # ✅ Get all screenshots related to the filtered attendance records
    screenshots = Screenshot.objects.filter(teacher_attendance__in=attendance_records)

    return render(request, "teacher_attendance/view_screenshot_per_date.html", {
        "attendance_records": attendance_records,
        "screenshots": screenshots,
        "selected_date": selected_date,
        "subject_id": subject_id,
    })


@login_required
def export_screenshots_pdf(request, subject_id, selected_date):
    """
    Generate a PDF with all screenshots and timestamps for the selected date.
    """
    try:
        selected_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Get subject and teacher information
    subject = get_object_or_404(Subject, id=subject_id)
    attendance_records = Teacher_Attendance.objects.filter(subject_id=subject_id, time_started__date=selected_date)
    
    # Get the teacher name (assuming the first attendance record is valid)
    teacher_name = attendance_records.first().teacher.get_full_name() if attendance_records.exists() else "Unknown"

    # Get all screenshots
    screenshots = Screenshot.objects.filter(teacher_attendance__in=attendance_records)

    # Create a PDF response
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="screenshots_{selected_date}.pdf"'

    # Create a PDF object
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter  # Get the PDF page size

    # ✅ **Title & Header**
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, height - 50, f"Screenshots for {selected_date}")

    pdf.setFont("Helvetica", 12)
    pdf.drawString(50, height - 70, f"Subject: {subject.subject_name}")
    pdf.drawString(50, height - 90, f"Teacher: {teacher_name}")

    y_position = height - 140  # Start positioning images after header

    for screenshot in screenshots:
        if y_position < 250:  # Avoid overlapping, start a new page
            pdf.showPage()
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(50, height - 50, f"Screenshots for {selected_date}")
            pdf.setFont("Helvetica", 12)
            pdf.drawString(50, height - 70, f"Subject: {subject.subject_name}")
            pdf.drawString(50, height - 90, f"Teacher: {teacher_name}")
            y_position = height - 140  # Reset position for new page

        # ✅ **Draw timestamp above the image**
        pdf.setFont("Helvetica", 12)
        pdf.drawString(50, y_position, f"Taken on: {screenshot.timestamp.strftime('%Y-%m-%d %H:%M:%S')}")

        # ✅ **Load and fit image properly**
        try:
            image = ImageReader(screenshot.image.path)
            img_width = width - 100  # Adjust width for margins
            img_height = 200  # Fixed height for uniformity
            pdf.drawImage(image, 50, y_position - img_height - 10, width=img_width, height=img_height, preserveAspectRatio=True)
            y_position -= img_height + 40  # Move down after image & timestamp
        except Exception as e:
            pdf.drawString(50, y_position - 20, "Error loading image")
            y_position -= 40

    # Save PDF
    pdf.save()
    buffer.seek(0)
    response.write(buffer.read())
    return response

@csrf_exempt
def save_screenshot(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data"}, status=400)

        image_data = data.get("image")
        attendance_id = data.get("attendance_id")

        if not image_data or not attendance_id:
            return JsonResponse({"error": "Missing image data or attendance_id"}, status=400)

        # Ensure attendance exists
        try:
            attendance = Teacher_Attendance.objects.get(id=attendance_id)
        except Teacher_Attendance.DoesNotExist:
            return JsonResponse({"error": "Attendance record not found"}, status=404)

        # Fix base64 decoding: Remove data URI prefix
        if "," in image_data:
            format_info, imgstr = image_data.split(";base64,")
        else:
            return JsonResponse({"error": "Invalid base64 format"}, status=400)

        ext = format_info.split("/")[-1]  # Get file extension (png/jpg)

        # Decode and save image
        try:
            image_file = ContentFile(base64.b64decode(imgstr), name=f"screenshot_{now().strftime('%Y%m%d_%H%M%S')}.{ext}")
            screenshot = Screenshot.objects.create(teacher_attendance=attendance, image=image_file)
        except Exception as e:
            return JsonResponse({"error": f"Failed to decode image: {str(e)}"}, status=400)

        return JsonResponse({"message": "Screenshot saved successfully", "image_url": screenshot.image.url})

    return JsonResponse({"error": "Invalid request"}, status=400)


def get_matching_dates(schedule_days, start_date, end_date):
    """
    Returns a list of dates within the range [start_date, end_date] that match the schedule's days_of_week.
    """
    days_map = {
        "Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3, "Fri": 4, "Sat": 5, "Sun": 6
    }
    matching_dates = []
    if not start_date or not end_date:
        return matching_dates

    # Convert schedule_days to actual weekday numbers
    schedule_day_numbers = [days_map[day] for day in schedule_days if day in days_map]

    # Iterate through all days in the range
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() in schedule_day_numbers:
            matching_dates.append(current_date)
        current_date += timedelta(days=1)

    return matching_dates



@login_required
def teacher_attendance_report(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # Convert start_date and end_date to datetime format
    try:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Get all subjects and assigned teachers
    subjects = Subject.objects.values(
        "id", "subject_name", "assign_teacher__id",
        "assign_teacher__first_name", "assign_teacher__last_name"
    )

    # Subquery to get total attendance per subject
    attendance_subquery = Teacher_Attendance.objects.filter(
        subject=OuterRef("id"),
        time_started__isnull=False,
        time_ended__isnull=False
    )

    if start_date:
        attendance_subquery = attendance_subquery.filter(time_started__date__gte=start_date)
    if end_date:
        attendance_subquery = attendance_subquery.filter(time_started__date__lte=end_date)

    attendance_subquery = (
        attendance_subquery
        .values("subject")
        .annotate(total_attendance=Sum(ExpressionWrapper(
            F("time_ended") - F("time_started"), output_field=DurationField()
        )))
        .values("total_attendance")
    )

    # Attach attendance time to subjects using Coalesce (default to 0 if missing)
    subjects = subjects.annotate(
        total_attendance_time=Coalesce(Subquery(attendance_subquery),Value(timedelta(0), output_field=DurationField()),output_field=DurationField()
        )
    )

    # Process schedule data
    schedule_data = Schedule.objects.all()
    
    schedule_lookup = {}
    for schedule in schedule_data:
        key = (
            schedule.subject.id, schedule.subject.subject_name,
            schedule.subject.assign_teacher.first_name, schedule.subject.assign_teacher.last_name
        )

        schedule_days = schedule.days_of_week  # e.g., ["Tue", "Thu"]
        if key not in schedule_lookup:
            schedule_lookup[key] = timedelta()

        # Find matching dates in the selected range
        matching_dates = get_matching_dates(schedule_days, start_date, end_date)

        total_scheduled_time = timedelta()  # ✅ Reset for each subject
        for date in matching_dates:
            schedule_start = datetime.combine(date, schedule.schedule_start_time)
            schedule_end = datetime.combine(date, schedule.schedule_end_time)
            total_scheduled_time += schedule_end - schedule_start  # ✅ Correct calculation

        schedule_lookup[key] = total_scheduled_time

    # Convert data for rendering
    report_data = []
    for record in subjects:
        key = (record["id"], record["subject_name"], record["assign_teacher__first_name"], record["assign_teacher__last_name"])
        scheduled_time = schedule_lookup.get(key, timedelta())  # Get scheduled time, default to 0

        attendance_without_seconds = timedelta(
            hours=record["total_attendance_time"].seconds // 3600,
            minutes=(record["total_attendance_time"].seconds % 3600) // 60
        )
        
        variance = scheduled_time - attendance_without_seconds

        # Convert duration fields to "hh:mm" format
        def format_duration(duration):
            if isinstance(duration, int):  # If already in seconds
                total_seconds = duration
            elif isinstance(duration, timedelta):  # Normal timedelta case
                total_seconds = int(duration.total_seconds())
            else:
                return "0:00"

            hours, remainder = divmod(total_seconds, 3600)
            minutes, _ = divmod(remainder, 60)
            return f"{hours}:{minutes:02d}"

        report_data.append({
            "teacher_id": record["assign_teacher__id"],
            "subject_id": record["id"],
            "teacher_name": f"{record['assign_teacher__first_name']} {record['assign_teacher__last_name']}",
            "subject_name": record["subject_name"],
            "total_attendance_time": format_duration(record["total_attendance_time"]),
            "total_scheduled_time": format_duration(scheduled_time),
            "variance": format_duration(variance),
        })

    return render(request, 'teacher_attendance/teacher_attendance_report.html', {
        "report_data": report_data,
        "start_date": start_date.strftime("%Y-%m-%d") if start_date else "",
        "end_date": end_date.strftime("%Y-%m-%d") if end_date else "",
    })




@login_required
def teacher_attendance_calendar(request):
    subject_id = request.GET.get("subject_id")  # Get subject ID from request
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # Convert start_date and end_date to datetime format
    try:
        start_date = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
        end_date = datetime.strptime(end_date, "%Y-%m-%d") if end_date else None
    except ValueError:
        return JsonResponse({"error": "Invalid date format"}, status=400)

    attendance_data = Teacher_Attendance.objects.filter(
        time_started__date__lte=localtime().date(), 
        time_ended__isnull=False,
        subject_id=subject_id
    )

    if start_date:
        attendance_data = attendance_data.filter(time_started__date__gte=start_date)
    if end_date:
        attendance_data = attendance_data.filter(time_started__date__lte=end_date)

    attendance_data = (
        attendance_data
        .annotate(date=TruncDate('time_started'))  # Group by date
        .values('date', 'subject__id')
        .annotate(
            total_attendance_time=Sum(ExpressionWrapper(
                F('time_ended') - F('time_started'), output_field=DurationField()
            ))
        )
    )

    # Format data for FullCalendar
    events = []
    for record in attendance_data:
        total_seconds = int(record["total_attendance_time"].total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        formatted_time = f"{hours}h {minutes}m"

        events.append({
            "title": formatted_time,  # Only display time
            "start": record["date"].strftime('%Y-%m-%d'),
        })

    return JsonResponse(events, safe=False)



@login_required
def teacher_attendance_calendar_page(request, subject_id):
    # Get subject details
    subject = Schedule.objects.filter(subject__id=subject_id).values(
        'subject__id',
        'subject__subject_name', 
        'subject__assign_teacher__first_name', 
        'subject__assign_teacher__last_name'
    ).first()

    return render(request, 'teacher_attendance/teacher_attendance_calendar.html', {
        "subject": subject,
        "subject_id": subject_id
    })

